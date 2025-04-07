import sqlite3
import os
from openpyxl import load_workbook
import numpy as np
import shutil

def safe_float_conversion(value):
    try:
        return float(value)
    except (ValueError, TypeError):
        return 0.0

def encontrar_primeira_linha_vazia_grupo(sheet, col_indices, linha_inicial=1):
    """
    Encontra, para um grupo de colunas, a maior primeira linha vazia
    dentre elas. Isso garante que o preenchimento será alinhado.
    """
    maior_primeira_livre = 0
    for col_idx in col_indices:
        linha = linha_inicial
        while sheet.cell(row=linha, column=col_idx).value not in (None, ""):
            linha += 1
        # Mantém o maior índice de linha encontrado
        if linha > maior_primeira_livre:
            maior_primeira_livre = linha
    return maior_primeira_livre

def gerar_planilha_para_arquivos(arquivos_selecionados, tipo_ensaio_selecionado, metodo):
    # Determinar o modelo de planilha com base no TipoEnsaio
    if tipo_ensaio_selecionado.startswith('TIR'):
        modelo_planilha = r'C:\Users\lgv_v\Documents\LUIZ\Modelo Planilha Final\ModeloPlanilhaFinal_TIR.xlsx'
    elif tipo_ensaio_selecionado.startswith('TER'):
        modelo_planilha = r'C:\Users\lgv_v\Documents\LUIZ\Modelo Planilha Final\ModeloPlanilhaFinal_TER.xlsx'
    else:
        print(f"TipoEnsaio '{tipo_ensaio_selecionado}' não reconhecido.")
        return

    if not os.path.exists(modelo_planilha):
        print(f"O modelo de planilha não foi encontrado em: {modelo_planilha}")
        return

    # Copiar o modelo para criar a nova planilha
    novo_arquivo = os.path.join(
        r'C:\Users\lgv_v\Documents\LUIZ\Modelo Planilha Final',
        f'Planilha_Preenchida_{tipo_ensaio_selecionado}_{metodo}.xlsx'
    )
    shutil.copy(modelo_planilha, novo_arquivo)

    # Abrir o workbook
    wb = load_workbook(novo_arquivo)

    db_path = r'C:\Users\lgv_v\Documents\LUIZ\database.db'
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()

    if not arquivos_selecionados:
        print("Nenhum arquivo selecionado para gerar a planilha.")
        conn.close()
        return

    # Dicionário para mapear índices a letras (A, B, C, D, E)
    letras = ['A', 'B', 'C', 'D', 'E']

    for idx, arquivo in enumerate(arquivos_selecionados):
        if idx >= len(letras):
            print(f"Mais de {len(letras)} arquivos selecionados. Apenas os primeiros {len(letras)} serão processados.")
            break

        # Nome da folha correspondente
        sheet_name = f"CP {letras[idx]} Data"

        # Verificar se a folha já existe
        if sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
        else:
            print(f"A planilha não contém a aba '{sheet_name}'.")
            continue

        # Recuperar idensaio via Cp table
        cursor.execute("""
            SELECT idensaio FROM Cp WHERE filename = ?
        """, (arquivo,))
        result = cursor.fetchone()
        if not result:
            print(f"Não foi possível encontrar idensaio para o arquivo {arquivo}")
            continue
        idensaio = result[0]

        # Recuperar metadados
        cursor.execute("""
            SELECT * FROM MetadadosArquivo WHERE idnome = ?
        """, (idensaio,))
        metadados_row = cursor.fetchone()
        if not metadados_row:
            print(f"Não foi possível encontrar metadados para idnome {idensaio}")
            continue
        columns = [description[0] for description in cursor.description]
        metadados = dict(zip(columns, metadados_row))

        # Obter os estágios dos metadados
        try:
            B_stage = int(metadados['_B'])
            Adensamento_stage = int(metadados['_ad'])
            cis_inicial = int(metadados['_cis_inicial'])
            cis_final = int(metadados['_cis_final'])
        except KeyError as e:
            print(f"Estágio '{e.args[0]}' não encontrado nos metadados do arquivo {arquivo}.")
            continue
        except ValueError:
            print(f"Valores inválidos para os estágios nos metadados do arquivo {arquivo}.")
            continue

        # Recuperar dados de EnsaiosTriaxiais
        cursor.execute("""
            SELECT * FROM EnsaiosTriaxiais WHERE idnome = ?
        """, (idensaio,))
        ensaio_data = cursor.fetchall()
        if not ensaio_data:
            print(f"Nenhum dado encontrado para idensaio {idensaio}")
            continue
        colunas = [description[0] for description in cursor.description]

        data_dict = {col: [] for col in colunas}
        for row in ensaio_data:
            for col, value in zip(colunas, row):
                data_dict[col].append(safe_float_conversion(value))

        def get_stage_data(stage_number):
            indices = [i for i, val in enumerate(data_dict['stage_no']) if val == stage_number]
            if not indices:
                print(f"Nenhum dado encontrado para o estágio {stage_number} no arquivo {arquivo}")
                return {}
            stage_data = {col: [data_dict[col][i] for i in indices] for col in colunas}
            return stage_data

        # B_data e Adensamento_data para cada estágio exato
        B_data = get_stage_data(B_stage)
        adensamento_data = get_stage_data(Adensamento_stage)

        # Filtrar dados de cisalhamento entre cis_inicial e cis_final
        cis_indices = [i for i, val in enumerate(data_dict['stage_no']) if cis_inicial <= val <= cis_final]
        if not cis_indices:
            print(f"Nenhum dado de cisalhamento encontrado para os estágios {cis_inicial} a {cis_final} no arquivo {arquivo}")
            continue

        cis_data = {col: [data_dict[col][i] for i in cis_indices] for col in colunas}

        if not B_data or not adensamento_data or not cis_data:
            print(f"Dados incompletos para o arquivo {arquivo}")
            continue

        # ========== ESTÁGIO B ==========
        # Mapeamento: (coluna_no_data_dict, coluna_no_Excel)
        b_map = [
            ("time_stage_start", 1),   # A
            ("rad_press_Original", 3), # C
            ("back_press_Original", 4),# D
            ("pore_press_Original", 5) # E
        ]
        # Maior número de linhas (entre as colunas mapeadas)
        b_rows = max(len(B_data.get(col_python, [])) for (col_python, _) in b_map)

        # Encontra a primeira linha livre considerando as colunas de B
        b_col_indices = [col_excel for (_, col_excel) in b_map]
        b_start_row = encontrar_primeira_linha_vazia_grupo(sheet, b_col_indices, linha_inicial=10)

        for i in range(b_rows):
            for (col_python, col_excel) in b_map:
                col_list = B_data.get(col_python, [])
                valor = col_list[i] if i < len(col_list) else None
                sheet.cell(row=b_start_row + i, column=col_excel, value=valor)

        # ========== ESTÁGIO ADENSAMENTO ==========
        ad_columns = [
            'stage_no', 'time_test_start', 'time_stage_start',
            'rad_press_Original', 'rad_vol_Original', 'back_press_Original',
            'back_vol_Original', 'load_cell_Original', 'pore_press_Original', 'ax_disp_Original'
        ]
        # AL=38 até AU=47 (10 colunas)
        ad_column_indices = [col for col in range(38, 48)]  # 38..47

        ad_max_rows = max(len(adensamento_data.get(col, [])) for col in ad_columns)

        # Encontra a primeira linha livre considerando as colunas de Adensamento
        ad_start_row = encontrar_primeira_linha_vazia_grupo(sheet, ad_column_indices, linha_inicial=7)

        for idx_row in range(ad_max_rows):
            for idx_col, col_name in enumerate(ad_columns):
                col_list = adensamento_data.get(col_name, [])
                value = col_list[idx_row] if idx_row < len(col_list) else None
                sheet.cell(row=ad_start_row + idx_row, column=ad_column_indices[idx_col], value=value)

        # ========== ESTÁGIO CISALHAMENTO ==========

        # 1) pore_press_0
        pore_press_list = cis_data.get('pore_press_Original', [])
        pore_press_0 = pore_press_list[0] if pore_press_list else 0

        # 2) eff_rad_stress
        rad_press_list = cis_data.get('rad_press_Original', [])
        eff_rad_stress = [rad_p - pore_p for rad_p, pore_p in zip(rad_press_list, pore_press_list)]

        # 3) dev_stress_B
        dev_stress_B = cis_data.get('dev_stress_B', [])
        eff_ax_stress_B = [dev + eff_rad for dev, eff_rad in zip(dev_stress_B, eff_rad_stress)]

        # 4) stress_ratio
        stress_ratio = [
            ea / er if er != 0 else None
            for ea, er in zip(eff_ax_stress_B, eff_rad_stress)
        ]

        # 5) pore_press_diff
        pore_press_diff = [pp - pore_press_0 for pp in pore_press_list]

        # 6) eff_stress_avg e eff_stress_diff
        eff_stress_avg = [(ea + er) / 2 for ea, er in zip(eff_ax_stress_B, eff_rad_stress)]
        eff_stress_diff = [(ea - er) / 2 for ea, er in zip(eff_ax_stress_B, eff_rad_stress)]

        # 7) Ajustar back_vol_Original
        back_vol_list = cis_data.get('back_vol_Original', [])
        adjusted_back_vol_Original = []
        if back_vol_list:
            first_back_vol = back_vol_list[0]
            adjusted_back_vol_Original = [val - first_back_vol for val in back_vol_list]

        # 8) Ajustar ax_disp_Original
        ax_disp_list = cis_data.get('ax_disp_Original', [])
        adjusted_ax_disp_Original = []
        if ax_disp_list:
            first_ax_disp = ax_disp_list[0]
            adjusted_ax_disp_Original = [val - first_ax_disp for val in ax_disp_list]

        # 9) ax_strain_percent e corrected_deviator_stress
        ax_strain_percent = cis_data.get('ax_strain', [])
        corrected_deviator_stress = dev_stress_B  # = q

        # Organiza dados para preencher
        cis_data_preenchimento = {
            'time_stage_start': cis_data.get('time_stage_start', []),
            'time_stage_start_div_60': [t / 60 for t in cis_data.get('time_stage_start', [])],
            'rad_press_Original': rad_press_list,
            'back_press_Original': cis_data.get('back_press_Original', []),
            'pore_press_Original': pore_press_list,
            'adjusted_back_vol_Original': adjusted_back_vol_Original,
            'adjusted_ax_disp_Original': adjusted_ax_disp_Original,
            'load_cell_Original': cis_data.get('load_cell_Original', []),
            'ax_strain_percent': ax_strain_percent,
            'corrected_deviator_stress': corrected_deviator_stress,
            'stress_ratio': stress_ratio,
            'pore_press_diff': pore_press_diff,
            'eff_stress_avg': eff_stress_avg,
            'eff_stress_diff': eff_stress_diff
        }

        # Mapeamento p/ Excel (R=18 até AE=31)
        cis_column_names = [
            'time_stage_start', 'time_stage_start_div_60',
            'rad_press_Original', 'back_press_Original',
            'pore_press_Original', 'adjusted_back_vol_Original',
            'adjusted_ax_disp_Original', 'load_cell_Original',
            'ax_strain_percent', 'corrected_deviator_stress',
            'stress_ratio', 'pore_press_diff',
            'eff_stress_avg', 'eff_stress_diff'
        ]
        cis_column_indices = [col for col in range(18, 32)]  # 18..31 (R..AE)

        cis_max_rows = max(len(cis_data_preenchimento[n]) for n in cis_column_names)

        # Achar primeira linha livre para as colunas de cisalhamento
        cis_start_row = encontrar_primeira_linha_vazia_grupo(sheet, cis_column_indices, linha_inicial=7)

        # Preencher
        for i in range(cis_max_rows):
            for idx_col, col_name in enumerate(cis_column_names):
                col_vals = cis_data_preenchimento[col_name]
                val = col_vals[i] if i < len(col_vals) else None
                sheet.cell(row=cis_start_row + i, column=cis_column_indices[idx_col], value=val)

    # Salvar e fechar o workbook
    wb.save(novo_arquivo)
    wb.close()
    conn.close()
    print(f"Planilha gerada com sucesso: {novo_arquivo}")
